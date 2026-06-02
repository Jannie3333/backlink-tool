# -*- coding: utf-8 -*-
"""Excel 保存工具"""
from pathlib import Path
from datetime import datetime
from io import BytesIO

OUTPUT_DIR = Path(__file__).parent / "output"


def save_research_excel(data: list, competitor_url: str) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "外链列表"

    headers = ["#","域名","来源页面","DR","流量","锚文本","联系邮箱","状态","备注"]
    widths  = [5, 28, 50, 8, 10, 28, 32, 12, 20]
    fh   = PatternFill(start_color="C8F135", end_color="C8F135", fill_type="solid")
    fa   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    fntH = Font(bold=True, color="000000", size=11)
    fntL = Font(color="1155CC", underline="single")
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill = fh; cell.font = fntH
        cell.alignment = ctr; cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w

    for i, row in enumerate(data, 1):
        r = i + 1
        vals = [i, row.get("referring_domain",""), row.get("referring_page",""),
                row.get("dr",""), row.get("traffic",""), row.get("anchor_text",""),
                row.get("email",""), "待联系", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = bdr
            cell.alignment = Alignment(vertical="center", wrap_text=(c==3))
            if i % 2 == 0: cell.fill = fa
            if c == 3 and v: cell.font = fntL

    ws.freeze_panes = "A2"
    clean = competitor_url.replace(".", "_").replace("/", "_")[:30]
    fname = f"backlinks_{clean}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    path  = OUTPUT_DIR / fname
    wb.save(path)
    return path


def save_submit_report(results: list) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "提交报告"

    headers = ["#","域名","结果","备注","提交时间"]
    widths  = [5, 35, 15, 45, 20]
    fh   = PatternFill(start_color="C8F135", end_color="C8F135", fill_type="solid")
    fntH = Font(bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill = fh; cell.font = fntH
        cell.alignment = Alignment(horizontal="center"); cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w

    STATUS_MAP = {
        "success":        "成功",
        "no_entry":       "无提交入口",
        "login_required": "需要登录",
        "error":          "失败",
        "skipped":        "跳过",
    }
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, r in enumerate(results, 1):
        for c, v in enumerate([
            i, r.get("domain",""),
            STATUS_MAP.get(r.get("status",""), r.get("status","")),
            r.get("note",""), ts
        ], 1):
            ws.cell(i+1, c, v).border = bdr

    ws.freeze_panes = "A2"
    fname = f"submit_report_{datetime.now():%Y%m%d_%H%M}.xlsx"
    path  = OUTPUT_DIR / fname
    wb.save(path)
    return path
