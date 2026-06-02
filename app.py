# -*- coding: utf-8 -*-
"""
外链工具箱 — Pixocto.ai
双功能：外链调研 + 外链自动提交
"""
import streamlit as st
import threading
import queue
import time
import pandas as pd
from io import BytesIO
from datetime import datetime
from pathlib import Path

from config import PRODUCT, IMAGES_DIR
from scraper import run_scraper
from submitter import run_submissions

# ─── 页面配置 ─────────────────────────────────────────────
st.set_page_config(
    page_title="外链工具箱",
    page_icon="🔗",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── 全局样式 ─────────────────────────────────────────────
st.markdown("""
<style>
  .stApp { background: #0f0f0f; color: #f0f0f0; }
  section[data-testid="stSidebar"] { background: #1a1a1a; }
  header[data-testid="stHeader"] { background: transparent; }

  .brand-bar {
    display: flex; align-items: center; gap: 12px;
    padding: 18px 0 6px 0; border-bottom: 1px solid #222; margin-bottom: 24px;
  }
  .brand-bar .logo { font-size: 20px; font-weight: 800; color: #c8f135; }
  .brand-bar .sub  { font-size: 12px; color: #555; }

  /* 模式卡片 */
  .mode-cards { display: flex; gap: 16px; margin-bottom: 28px; }
  .mode-card {
    flex: 1; background: #1a1a1a; border: 2px solid #2a2a2a;
    border-radius: 14px; padding: 22px 24px; cursor: pointer;
    transition: all .2s;
  }
  .mode-card.active { border-color: #c8f135; background: #1f2010; }
  .mode-card .icon  { font-size: 28px; margin-bottom: 10px; }
  .mode-card .title { font-size: 15px; font-weight: 700; color: #fff; margin-bottom: 6px; }
  .mode-card .desc  { font-size: 12px; color: #666; line-height: 1.5; }
  .mode-card.active .title { color: #c8f135; }

  /* Flow arrow */
  .flow-arrow {
    display: flex; align-items: center; justify-content: center;
    color: #c8f135; font-size: 22px; font-weight: 700; margin: 0 4px;
  }

  /* Tabs */
  .stTabs [data-baseweb="tab-list"] { background: transparent; gap: 6px; }
  .stTabs [data-baseweb="tab"] {
    background: #1e1e1e; color: #888; border-radius: 8px;
    padding: 8px 22px; font-weight: 600; border: 1px solid #2a2a2a;
  }
  .stTabs [aria-selected="true"] {
    background: #c8f135 !important; color: #000 !important; border-color: #c8f135 !important;
  }

  /* Inputs */
  .stTextInput input, .stTextArea textarea {
    background: #1e1e1e !important; color: #fff !important;
    border: 1px solid #333 !important; border-radius: 8px !important;
  }

  /* Buttons */
  .stButton > button {
    background: #c8f135; color: #000; font-weight: 700;
    border: none; border-radius: 8px; padding: 10px 26px; transition: all .15s;
  }
  .stButton > button:hover { background: #d6ff3e; transform: translateY(-1px); }

  /* 次要按钮 */
  button[kind="secondary"] {
    background: #1e1e1e !important; color: #c8f135 !important;
    border: 1px solid #c8f135 !important;
  }

  /* Import banner */
  .import-banner {
    background: #1f2010; border: 1px solid #c8f135; border-radius: 10px;
    padding: 14px 18px; display: flex; align-items: center; gap: 12px;
    margin-bottom: 16px;
  }
  .import-banner .ib-icon { font-size: 20px; }
  .import-banner .ib-text { font-size: 13px; color: #c8f135; font-weight: 600; }
  .import-banner .ib-sub  { font-size: 11px; color: #888; }

  /* Log box */
  .log-box {
    background: #0d1a0d; border: 1px solid #1a3a1a; border-radius: 10px;
    padding: 16px; font-family: monospace; font-size: 12.5px;
    color: #7fff7f; max-height: 400px; overflow-y: auto;
    white-space: pre-wrap; line-height: 1.75;
  }

  /* Stat card */
  .stat-card {
    background: #1a1a1a; border: 1px solid #2a2a2a; border-radius: 12px;
    padding: 16px 20px; text-align: center;
  }
  .stat-card .num { font-size: 30px; font-weight: 800; color: #c8f135; }
  .stat-card .lbl { font-size: 11px; color: #666; margin-top: 4px; }

  hr { border-color: #222; }
  .stDataFrame { background: #1a1a1a; }

  /* Step badge */
  .step-badge {
    display: inline-block; background: #c8f135; color: #000;
    font-weight: 800; font-size: 11px; border-radius: 20px;
    padding: 2px 10px; margin-right: 8px;
  }
</style>
""", unsafe_allow_html=True)

# ─── 品牌栏 ───────────────────────────────────────────────
st.markdown("""
<div class="brand-bar">
  <span class="logo">🔗 外链工具箱</span>
  <span class="sub">Pixocto.ai · 外链调研 & 自动提交</span>
</div>
""", unsafe_allow_html=True)

# ─── Session State 初始化 ─────────────────────────────────
DEFAULTS = {
    "mode":              None,           # "full" | "direct"
    # 调研
    "research_logs":     [],
    "research_running":  False,
    "research_results":  None,
    "research_queue":    None,
    # 提交
    "submit_logs":       [],
    "submit_running":    False,
    "submit_results":    None,
    "submit_queue":      None,
    "submit_done":       False,
    "all_results":       [],
    "batch_count":       0,
    # 从调研导入的域名
    "imported_domains":  [],
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─── 工具函数 ─────────────────────────────────────────────
def _make_excel(data: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "外链列表"
    headers = ["#","域名","来源页面","DR","流量","锚文本","联系邮箱","状态","备注"]
    widths  = [5,28,48,8,10,28,32,12,20]
    fh = PatternFill(start_color="C8F135",end_color="C8F135",fill_type="solid")
    fa = PatternFill(start_color="1A1A1A",end_color="1A1A1A",fill_type="solid")
    fnt_h = Font(bold=True,color="000000",size=11)
    fnt_l = Font(color="4FC3F7",underline="single")
    thin  = Side(style="thin",color="333333")
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.row_dimensions[1].height = 26
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(1,c,h); cell.fill=fh; cell.font=fnt_h
        cell.alignment=Alignment(horizontal="center"); cell.border=bdr
        ws.column_dimensions[cell.column_letter].width=w
    for i,row in enumerate(data,1):
        r=i+1
        vals=[i,row.get("referring_domain",""),row.get("referring_page",""),
              row.get("dr",""),row.get("traffic",""),row.get("anchor_text",""),
              row.get("email",""),"待联系",""]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.border=bdr
            cell.alignment=Alignment(vertical="center",wrap_text=(c==3))
            if i%2==0: cell.fill=fa
            if c==3 and v: cell.font=fnt_l
    ws.freeze_panes="A2"
    buf=BytesIO(); wb.save(buf); return buf.getvalue()


def _make_report_excel(results: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "提交报告"
    headers=["#","域名","结果","备注","提交时间"]; widths=[5,35,15,45,20]
    fh=PatternFill(start_color="C8F135",end_color="C8F135",fill_type="solid")
    fnt_h=Font(bold=True,size=11); thin=Side(style="thin",color="333333")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(1,c,h); cell.fill=fh; cell.font=fnt_h
        cell.alignment=Alignment(horizontal="center"); cell.border=bdr
        ws.column_dimensions[cell.column_letter].width=w
    STATUS_MAP={"success":"✅ 成功","no_entry":"⚠️ 无入口",
                "login_required":"🔑 需登录","error":"❌ 失败","skipped":"⏭️ 跳过"}
    ts=datetime.now().strftime("%Y-%m-%d %H:%M")
    for i,r in enumerate(results,1):
        for c,v in enumerate([i,r.get("domain",""),
                               STATUS_MAP.get(r.get("status",""),r.get("status","")),
                               r.get("note",""),ts],1):
            ws.cell(i+1,c,v).border=bdr
    ws.freeze_panes="A2"
    buf=BytesIO(); wb.save(buf); return buf.getvalue()


def _start_submit_thread(domains: list):
    """启动提交后台线程"""
    st.session_state.submit_logs    = []
    st.session_state.submit_results = None
    st.session_state.submit_done    = False
    st.session_state.submit_running = True
    q = queue.Queue()
    st.session_state.submit_queue   = q

    def _worker(doms, q):
        def cb(msg): q.put(("log", msg))
        try:
            res = run_submissions(doms, log_cb=cb)
            q.put(("done", res))
        except Exception as e:
            q.put(("error", str(e)))

    threading.Thread(target=_worker, args=(domains, q), daemon=True).start()


def _submit_log_panel():
    """消费提交队列并渲染日志，返回 True 表示仍在运行"""
    q = st.session_state.submit_queue
    if q:
        while not q.empty():
            t, payload = q.get_nowait()
            if t == "log":
                st.session_state.submit_logs.append(payload)
            elif t == "done":
                st.session_state.submit_results  = payload
                st.session_state.all_results.extend(payload)
                st.session_state.batch_count    += 1
                st.session_state.submit_running  = False
                st.session_state.submit_done     = True
            elif t == "error":
                st.session_state.submit_logs.append(f"❌ 错误: {payload}")
                st.session_state.submit_running  = False
                st.session_state.submit_done     = True

    if st.session_state.submit_logs:
        log_html = "\n".join(st.session_state.submit_logs[-120:])
        st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)

    return st.session_state.submit_running


def _submit_done_panel():
    """批次完成后的「继续/停止」面板"""
    res  = st.session_state.submit_results or []
    ok   = sum(1 for r in res if r["status"] == "success")
    fail = len(res) - ok
    tot  = len(st.session_state.all_results)

    st.divider()
    st.markdown(f"**本批完成：** ✅ {ok} 成功 &nbsp;｜&nbsp; ❌ {fail} 失败 &nbsp;｜&nbsp; 累计已提交 **{tot}** 个站")

    c1, c2, _ = st.columns([2, 2, 4])
    with c1:
        cont = st.button("➕ 继续下一批", key="btn_cont")
    with c2:
        stop = st.button("🛑 停止 & 生成报告", key="btn_stop", type="secondary")

    if cont:
        st.session_state.submit_done    = False
        st.session_state.submit_logs    = []
        st.session_state.submit_results = None
        st.session_state.submit_queue   = None
        st.rerun()

    if stop:
        _show_final_report()


def _show_final_report():
    all_res = st.session_state.all_results
    ok  = sum(1 for r in all_res if r["status"] == "success")
    fail= len(all_res) - ok

    st.divider()
    st.markdown("## 📊 外链提交完整报告")
    c1,c2,c3,c4 = st.columns(4)
    for col, num, lbl in [
        (c1, len(all_res),                    "总提交"),
        (c2, ok,                               "✅ 成功"),
        (c3, fail,                             "❌ 失败/跳过"),
        (c4, st.session_state.batch_count,    "批次数"),
    ]:
        col.markdown(f'<div class="stat-card"><div class="num">{num}</div>'
                     f'<div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    df = pd.DataFrame(all_res)
    st.dataframe(df, use_container_width=True, height=360)

    st.download_button(
        "⬇️ 下载完整报告 Excel",
        data=_make_report_excel(all_res),
        file_name=f"submit_report_{datetime.now():%Y%m%d_%H%M}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if st.button("🔄 重新开始", key="btn_reset"):
        for k, v in DEFAULTS.items():
            st.session_state[k] = v
        st.rerun()


# ═══════════════════════════════════════════════════════════
#  模式选择（首屏）
# ═══════════════════════════════════════════════════════════
if st.session_state.mode is None:
    st.markdown("### 选择工作模式")
    st.markdown("<br>", unsafe_allow_html=True)

    col_a, col_arrow, col_b = st.columns([5, 1, 5])

    with col_a:
        st.markdown("""
        <div class="mode-card">
          <div class="icon">🔍 → 📤</div>
          <div class="title">完整流程</div>
          <div class="desc">
            先用 Ahrefs 查询竞品外链，<br>
            系统自动收集联系邮箱，<br>
            然后一键把结果导入提交队列，<br>
            按顺序完成调研 + 提交。
          </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("选择完整流程", key="mode_full", use_container_width=True):
            st.session_state.mode = "full"
            st.rerun()

    with col_arrow:
        st.markdown('<div style="height:120px;display:flex;align-items:center;'
                    'justify-content:center;color:#444;font-size:24px;">或</div>',
                    unsafe_allow_html=True)

    with col_b:
        st.markdown("""
        <div class="mode-card">
          <div class="icon">📤</div>
          <div class="title">直接提交</div>
          <div class="desc">
            跳过调研，<br>
            直接粘贴 20 个域名，<br>
            工具自动逐一访问并提交外链，<br>
            完成后可继续下一批或导出报告。
          </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("选择直接提交", key="mode_direct", use_container_width=True):
            st.session_state.mode = "direct"
            st.rerun()

    st.stop()


# ─── 顶部：已选模式 + 切换按钮 ───────────────────────────
mode_label = "🔍 → 📤  完整流程" if st.session_state.mode == "full" else "📤  直接提交"
col_ml, col_mb = st.columns([7, 1])
col_ml.markdown(f"**当前模式：** {mode_label}")
if col_mb.button("切换模式", key="switch_mode"):
    for k, v in DEFAULTS.items():
        st.session_state[k] = v
    st.rerun()

st.markdown("---")


# ═══════════════════════════════════════════════════════════
#  模式 A：完整流程（Tab 1 调研 → Tab 2 提交）
# ═══════════════════════════════════════════════════════════
if st.session_state.mode == "full":

    tab1, tab2 = st.tabs([
        "① 🔍 外链调研",
        "② 📤 外链提交",
    ])

    # ── Tab 1：调研 ──────────────────────────────────────
    with tab1:
        st.markdown('<span class="step-badge">STEP 1</span> 输入竞品域名，调研外链并收集邮箱', unsafe_allow_html=True)
        st.markdown("")

        c1, c2 = st.columns([5, 2])
        with c1:
            competitor = st.text_input("竞品域名", value="runwayml.com",
                                       placeholder="如 runway.com、pika.art",
                                       label_visibility="collapsed")
        with c2:
            max_links = st.number_input("最多外链数", 5, 50, 20)

        if st.button("🔍 开始调研", key="btn_research") and not st.session_state.research_running:
            st.session_state.research_logs   = []
            st.session_state.research_results= None
            st.session_state.research_running= True
            st.session_state.imported_domains= []
            q = queue.Queue(); st.session_state.research_queue = q

            def _r_worker(url, mx, q):
                def cb(m): q.put(("log", m))
                try:    q.put(("done", run_scraper(url, log_cb=cb, max_sites=mx)))
                except Exception as e: q.put(("error", str(e)))

            threading.Thread(target=_r_worker,
                             args=(competitor, max_links, st.session_state.research_queue),
                             daemon=True).start()

        # 日志
        rq = st.session_state.research_queue
        if rq and st.session_state.research_running:
            while not rq.empty():
                t, p = rq.get_nowait()
                if t == "log":  st.session_state.research_logs.append(p)
                elif t == "done":
                    st.session_state.research_results = p
                    st.session_state.research_running = False
                elif t == "error":
                    st.session_state.research_logs.append(f"❌ {p}")
                    st.session_state.research_running = False

        if st.session_state.research_logs:
            log_html = "\n".join(st.session_state.research_logs[-80:])
            st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)

        if st.session_state.research_running:
            time.sleep(0.8); st.rerun()

        # 结果
        if st.session_state.research_results is not None:
            data = st.session_state.research_results
            st.divider()

            c1,c2,c3 = st.columns(3)
            c1.markdown(f'<div class="stat-card"><div class="num">{len(data)}</div><div class="lbl">外链总数</div></div>', unsafe_allow_html=True)
            c2.markdown(f'<div class="stat-card"><div class="num">{sum(1 for r in data if r.get("email"))}</div><div class="lbl">找到邮箱</div></div>', unsafe_allow_html=True)
            c3.markdown(f'<div class="stat-card"><div class="num">{sum(1 for r in data if not r.get("email"))}</div><div class="lbl">未找到邮箱</div></div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            df = pd.DataFrame(data)[["referring_domain","referring_page","dr","traffic","anchor_text","email"]]
            df.columns = ["域名","来源页面","DR","流量","锚文本","联系邮箱"]
            st.dataframe(df, use_container_width=True, height=320)

            cola, colb = st.columns([2, 3])
            with cola:
                excel_bytes = _make_excel(data)
                st.download_button("⬇️ 下载 Excel",
                    data=excel_bytes,
                    file_name=f"backlinks_{competitor.replace('.','_')}_{datetime.now():%Y%m%d_%H%M}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with colb:
                if st.button("📤 导入到外链提交 →", key="btn_import"):
                    # 把调研域名写入 session，切到 Tab 2
                    st.session_state.imported_domains = [
                        r["referring_domain"] for r in data if r.get("referring_domain")
                    ]
                    st.toast(f"✅ 已导入 {len(st.session_state.imported_domains)} 个域名，请切换到 Tab ②", icon="📤")

    # ── Tab 2：提交 ──────────────────────────────────────
    with tab2:
        st.markdown('<span class="step-badge">STEP 2</span> 提交外链到目录站', unsafe_allow_html=True)
        st.markdown("")

        # 已导入提示
        imported = st.session_state.imported_domains
        if imported:
            st.markdown(f"""
            <div class="import-banner">
              <span class="ib-icon">📋</span>
              <div>
                <div class="ib-text">已从调研结果导入 {len(imported)} 个域名</div>
                <div class="ib-sub">将按每批 20 个自动分批提交</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # 分批展示
            batch_size = 20
            batches = [imported[i:i+batch_size] for i in range(0, len(imported), batch_size)]
            current_batch_idx = st.session_state.batch_count % len(batches) if batches else 0
            current_domains = batches[min(current_batch_idx, len(batches)-1)]

            st.caption(f"当前批次 {min(current_batch_idx+1, len(batches))} / {len(batches)}（共 {len(current_domains)} 个）：")
            st.code("\n".join(current_domains))

            if not st.session_state.submit_running and not st.session_state.submit_done:
                if st.button(f"📤 开始提交第 {current_batch_idx+1} 批", key="btn_submit_imported"):
                    _start_submit_thread(current_domains)
                    st.rerun()
        else:
            st.info("请先在 Tab ① 完成调研，然后点击「导入到外链提交」按钮。\n\n或者切换到「直接提交」模式手动输入域名。")

        # 日志 + 完成面板
        if st.session_state.submit_running:
            still_running = _submit_log_panel()
            if still_running:
                time.sleep(0.8); st.rerun()
        elif st.session_state.submit_done:
            _submit_log_panel()
            _submit_done_panel()


# ═══════════════════════════════════════════════════════════
#  模式 B：直接提交
# ═══════════════════════════════════════════════════════════
elif st.session_state.mode == "direct":

    st.markdown('<span class="step-badge">直接提交</span> 粘贴域名，立即开始', unsafe_allow_html=True)
    st.markdown("")

    # 产品信息折叠
    with st.expander("📋 产品信息（Pixocto.ai）", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"**名称：** {PRODUCT['name']}")
            st.markdown(f"**网址：** {PRODUCT['url']}")
            st.markdown(f"**价格：** {PRODUCT['price']}")
            st.markdown(f"**提交邮箱：** {PRODUCT['submitter_email']}")
        with c2:
            st.markdown(f"**提交人：** {PRODUCT['submitter_name']}")
            st.markdown(f"**Twitter：** {PRODUCT['social_twitter']}")
            st.markdown(f"**关键词：** {', '.join(PRODUCT['keywords'])}")

        st.markdown("---")
        st.markdown("**产品截图**（提交表单时自动上传）")
        uploaded = st.file_uploader("上传截图", type=["png","jpg","jpeg"],
                                    accept_multiple_files=True, key="img_up")
        if uploaded:
            IMAGES_DIR.mkdir(exist_ok=True)
            for f in uploaded:
                (IMAGES_DIR / f.name).write_bytes(f.read())
            st.success(f"已保存 {len(uploaded)} 张图片")
        imgs = list(IMAGES_DIR.glob("*.png")) + list(IMAGES_DIR.glob("*.jpg"))
        if imgs:
            st.caption(f"已有 {len(imgs)} 张：{', '.join(i.name for i in imgs[:5])}")

    st.markdown("---")

    if not st.session_state.submit_running and not st.session_state.submit_done:
        st.markdown("**粘贴域名（每行一个，最多 20 个）：**")
        domains_raw = st.text_area("", height=280,
                                   placeholder="futurepedia.io\ntheresanaiforthat.com\ntoolfinder.xyz\n...",
                                   label_visibility="collapsed")

        batch_num = st.session_state.batch_count + 1
        if st.button(f"📤 开始提交第 {batch_num} 批", key="btn_direct_submit"):
            domains = [d.strip().replace("https://","").replace("http://","").strip("/")
                       for d in domains_raw.strip().splitlines() if d.strip()][:20]
            if not domains:
                st.warning("请先输入至少一个域名")
            else:
                _start_submit_thread(domains)
                st.rerun()

    # 日志 + 完成面板
    if st.session_state.submit_running:
        still_running = _submit_log_panel()
        if still_running:
            time.sleep(0.8); st.rerun()
    elif st.session_state.submit_done:
        _submit_log_panel()
        _submit_done_panel()
